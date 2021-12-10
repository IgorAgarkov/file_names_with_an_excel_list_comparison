# -*- coding: utf8 -*-

from openpyxl import load_workbook
import os

files = os.listdir(r'D:\Рабочая\Заказы\Заявка №428 2021.08.20\Чертежи')  # ссылка на папку с CDW
files_no_cdw = []

# наполняем список files_no_cdw названиями файлов cdw без расширений
for i in files:
    if i[-4:].lower() == '.cdw':
        files_no_cdw.append(i[:-4])
wb = load_workbook(r'D:\Рабочая\Заказы\Заявка №428 2021.08.20\Заявка №428 от 20.08.2021.xlsx')  # ссылка на файл excel
ws = wb.active


excel_list = []

# наполняем список excel_list значениями ячеек
cell = 6  # номер начальной ячейки
while ws['B' + str(cell)].value != None:
    excel_list.append(ws['B' + str(cell)].value)
    cell += 1

flag = True

# сравниваем каждый файл со значением ячеек столбца B
for cdw in files_no_cdw:          # перебираем все файлы
    # c = ws['B' + str(cell)].value
    if cdw not in excel_list:
        print('Нет записи:', cdw)
        flag = False

print()

# сравниваем каждую ячейку (из диапазона) столбца B со списком файлов files_no_dxf
for exl in excel_list:          # диапазон строк в файле Excel
    # c = ws['B' + str(cell)].value
    if exl not in files_no_cdw:
        print('Нет файла:', exl)
        flag = False


if flag == True:
    print('Всё в порядке! Список позиций Excel и файлы DXF соответствуют друг другу.')
