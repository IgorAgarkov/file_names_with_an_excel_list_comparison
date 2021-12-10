# -*- coding: utf8 -*-

from openpyxl import load_workbook
# import os

wb = load_workbook('Заявка №593 2021.11.08.xlsx')  # ссылка на 1-й файл excel (больший)
ws = wb.active

wb_2 = load_workbook('дополнение к заявке №593.xlsx')  # ссылка на 2-й файл excel (вычитаемый)
ws_2 = wb_2.active

excel_dict = {}

# наполняем словарь excel_dict значениями ячеек из первого файла excel
cell = 6  # номер начальной ячейки
while ws['B' + str(cell)].value != None:
    excel_dict[ws['B' + str(cell)].value] = eval(ws['C' + str(cell)].value[1:])
    cell += 1

cell = 6  # номер начальной ячейки
while ws_2['B' + str(cell)].value != None:
    name = ws_2['B' + str(cell)].value
    if name in excel_dict and excel_dict[name] - eval(ws_2['C' + str(cell)].value[1:]) > 0:
        # print(f"{name}: +{excel_dict[name] - eval(ws_2['C' + str(cell)].value[1:])} к имеющимуся значению")
        print(f"{name}: заменить значение на {excel_dict[name]}")
    cell += 1
