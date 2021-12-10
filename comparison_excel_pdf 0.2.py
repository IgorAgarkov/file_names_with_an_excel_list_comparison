# -*- coding: utf8 -*-

from openpyxl import load_workbook
import os

files = os.listdir(r'D:\Рабочая\от Ивана__\клапана\КО-7\для чертежей_')  # ссылка на папку с DXF
files_no_dxf = []

# наполняем список files_no_dxf названиями файлов pdf без расширений
for i in files:
    if i[-4:].lower() == '.pdf':
        files_no_dxf.append(i[:-4])
wb = load_workbook(r'D:\Рабочая\Заказы\Заявка №458 2021.09.01\Заявка КО-7 458 от 00.00.2021.xlsx')  # ссылка на файл excel
ws = wb.active


excel_list = []

# наполняем список excel_list значениями ячеек
cell = 6  # номер начальной ячейки
while ws['B' + str(cell)].value != None:
    excel_list.append(ws['B' + str(cell)].value)
    cell += 1

flag = True

# сравниваем каждый файл со значением ячеек столбца B
for dxf in files_no_dxf:          # перебираем все файлы
    # c = ws['B' + str(cell)].value
    if dxf not in excel_list:
        print('Нет записи:', dxf)
        flag = False

print()

# сравниваем каждую ячейку (из диапазона) столбца B со списком файлов files_no_dxf
for exl in excel_list:          # диапазон строк в файле Excel
    # c = ws['B' + str(cell)].value
    if exl not in files_no_dxf:
        print('Нет файла:', exl)
        flag = False


if flag == True:
    print('Всё в порядке! Список позиций Excel и файлы DXF соответствуют друг другу.')
