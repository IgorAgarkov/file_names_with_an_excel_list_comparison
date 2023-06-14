# -*- coding: utf8 -*-

from openpyxl import load_workbook
import os

# excel_link = r'D:\tmp\Новая папка (4)\Заявка №хх от 00.00.2021.xlsx'   # ссылка на файл excel
lile_link  = r'.'                                    # ссылка на папку с файлами
ext    = 'dxf' # расшерение файла
column = 'B'   # столбец в excel
cell   = 4     # номер начальной ячейки в excel

files = os.listdir(lile_link)   # получаем список всех файлов в директории 
no_ext_files = []

# наполняем список no_ext_files названиями файлов dxf без расширений
for i in files:
    if i[-4:].lower() == '.' + ext:
        no_ext_files.append(i[:-(len(ext) + 1)])  # 
    elif i[-5:].lower() == '.xlsx':
        excel_link = i

wb = load_workbook(excel_link)           
ws = wb.active

excel_list = []

# наполняем список excel_list значениями ячеек
while ws[column + str(cell)].value != None:
    excel_list.append(ws[column + str(cell)].value)
    cell += 1

flag = True

# сравниваем каждый файл со значением ячеек столбца B
for dxf in no_ext_files:          # перебираем все файлы
    if dxf not in excel_list:
        print('Нет записи:', dxf)
        flag = False

print()

# сравниваем каждую ячейку (из диапазона) столбца B со списком файлов no_ext_files
for exl in excel_list:          # диапазон строк в файле Excel
    if exl not in no_ext_files:
        print('Нет файла:', exl)
        flag = False

if flag == True:
    print(f'Всё в порядке! Список позиций Excel и файлы {ext} соответствуют друг другу.')

print()
input("Нажмите Enter для выхода")
